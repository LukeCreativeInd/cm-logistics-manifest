import streamlit as st
import pandas as pd
import math
import zipfile
from io import BytesIO
import re
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tempfile import NamedTemporaryFile
from pathlib import Path

NAN_LIKE = {"nan", "none", "null", ""}

def clean_cell(x: object) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    s = str(x).strip()
    return "" if s.lower() in NAN_LIKE else s

def to_clean_str(val: object) -> str:
    s = clean_cell(val)
    if s.startswith("'"):
        s = s[1:]
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s

def to_intish_str(val: object) -> str:
    s = to_clean_str(val)
    if re.fullmatch(r"\d+(\.\d+)?", s):
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except:
            pass
    return s

def format_phone(phone: object) -> str:
    p = clean_cell(phone).replace(" ", "").replace("+", "")
    if not p:
        return ""
    if p.startswith("61"):
        p = "0" + p[2:]
    elif p.startswith("4"):
        p = "0" + p
    return p

def run():
    st.markdown("### Clean Eats Manifest Generator")

    uploaded_file = st.file_uploader("Upload Clean Eats orders_export CSV file", type="csv")
    generate = st.button("Generate Clean Eats Manifests")

    if not (uploaded_file and generate):
        return

    orders_df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    orders_df = orders_df.map(clean_cell)
    orders_df.columns = orders_df.columns.str.strip()

    expected_cols = [
        "Notes","Tags","Shipping Phone","Shipping Street","Shipping City","Shipping Zip",
        "Shipping Province","Shipping Country","Shipping Name","Shipping Company","Email",
        "Name","Lineitem name","Lineitem quantity"
    ]
    for c in expected_cols:
        if c not in orders_df.columns:
            orders_df[c] = ""

    bundle_items = [
        "CARB LOVER'S FEAST","SUPER CHARGED CALORIES","FEED ME BEEF","GIVE ME CHICKEN",
        "I WON'T PAS(TA) ON THIS MEAL","THE MEGA PACK","MAKE YOUR OWN MEGA PACK",
        "CARB HATERS FEAST","UNDER CHARGED CALORIES","VEGGIE LOVERS PACK","Clean Eats Meal Plan"
    ]
    family_double_items = ["Family Mac and 3 Cheese Pasta Bake","Baked Family Lasagna"]

    manifest_rows = []
    for name, group in orders_df.groupby("Name", sort=False):
        order = group.iloc[0]
        total_qty = 0
        for _, row in group.iterrows():
            item = clean_cell(row["Lineitem name"])
            qty_raw = clean_cell(row["Lineitem quantity"])
            try:
                qty = int(float(qty_raw))
            except:
                qty = 0
            if any(bundle in item for bundle in bundle_items):
                continue
            elif item in family_double_items:
                total_qty += qty * 2
            else:
                total_qty += qty

        labels = math.ceil(total_qty / 24) if total_qty else 0

        state_map = {"VIC": "Victoria","NSW":"New South Wales","ACT":"Australian Capital Territory"}
        country_map = {"AU":"Australia"}
        raw_state = clean_cell(order["Shipping Province"])
        raw_country = clean_cell(order["Shipping Country"])
        state = state_map.get(raw_state, raw_state)
        country = country_map.get(raw_country, raw_country)

        tags_blob = clean_cell(order["Tags"])
        m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", tags_blob)
        delivery_date = m.group(1) if m else ""

        manifest_rows.append({
            "D.O. No.": to_clean_str(name),
            "Date": delivery_date,
            "Address 1": clean_cell(order["Shipping Street"]),
            "Address 2": clean_cell(order["Shipping City"]),
            "Postal Code": to_clean_str(order["Shipping Zip"]),
            "State": state,
            "Country": country,
            "Deliver to": (clean_cell(order.get("Shipping Company", "")) or clean_cell(order.get("Shipping Name", ""))),
            "Phone No.": to_clean_str(format_phone(order["Shipping Phone"])),
            "Time Window": "0600-1800",
            "Group": "Clean Eats Australia",
            "No. of Shipping Labels": labels,
            "Line Items": total_qty,
            "Email": clean_cell(order["Email"]),
            "Instructions": clean_cell(order["Notes"])
        })

    manifest_df = pd.DataFrame(manifest_rows)

    tag_series = orders_df.groupby("Name")["Tags"].agg(lambda s: " ".join(map(clean_cell, s)))
    def names_with(tag): return tag_series[tag_series.str.contains(tag, na=False, case=False)].index.tolist()

    cm_names = names_with("CM"); mc_names = names_with("MC"); cx_names = names_with("CX"); dk_names = names_with("DK")
    all_tagged_names = set(cm_names) | set(mc_names) | set(cx_names) | set(dk_names)

    cm_manifest = manifest_df[manifest_df["D.O. No."].isin(cm_names)]
    mc_manifest = manifest_df[manifest_df["D.O. No."].isin(mc_names)]
    cx_manifest = manifest_df[manifest_df["D.O. No."].isin(cx_names)]
    other_manifest = manifest_df[~manifest_df["D.O. No."].isin(all_tagged_names)]

    def company_or_name(group: pd.DataFrame) -> str:
        comp = clean_cell(group["Shipping Company"].iloc[0]) if len(group) else ""
        name = clean_cell(group["Shipping Name"].iloc[0]) if len(group) else ""
        return comp if comp else name

    if not mc_manifest.empty:
        # Future-proof (avoid GroupBy.apply behavior changes): prefer Shipping Company, else Shipping Name
        fb = orders_df.groupby("Name", sort=False)[["Shipping Company", "Shipping Name"]].first().fillna("")
        fallback = fb["Shipping Company"].where(fb["Shipping Company"].astype(str).str.strip() != "", fb["Shipping Name"]).to_dict()
        mc_manifest = mc_manifest.copy()
        mc_manifest["Deliver to"] = mc_manifest["D.O. No."].map(fallback).fillna("")

    output = BytesIO()
    with zipfile.ZipFile(output, "w") as zipf:
        def add_to_zip_excel(df, filename):
            if df.empty: return
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                df = df.copy()
                for col in ["Phone No.","Postal Code"]:
                    if col in df.columns: df[col] = df[col].astype(str)
                df.to_excel(writer, index=False, sheet_name='Manifest')
                wb = writer.book; ws = writer.sheets['Manifest']
                text_fmt = wb.add_format({'num_format': '@'})
                for col in ["Phone No.","Postal Code"]:
                    if col in df.columns:
                        idx = df.columns.get_loc(col)
                        ws.set_column(idx, idx, None, text_fmt)
            zipf.writestr(filename, buffer.getvalue())

        add_to_zip_excel(cm_manifest, "CM_Manifest.xlsx")
        add_to_zip_excel(mc_manifest, "MC_Manifest.xlsx")
        # CX Cold Xpress (populate template)
        if not cx_manifest.empty:
            template_path = Path(__file__).resolve().parent / "cx_manifest_template.xlsx"
            if not template_path.exists():
                template_path = Path("cx_manifest_template.xlsx")
            wb_cx = load_workbook(template_path)
            ws_cx = wb_cx["Sheet1"] if "Sheet1" in wb_cx.sheetnames else wb_cx.active

            today_mel = datetime.now(ZoneInfo("Australia/Melbourne")).date()
            cx_date_str = (today_mel + timedelta(days=1)).strftime("%d/%m/%Y")

            # Header cells
            ws_cx["B3"] = "Clean Eats Australia"
            ws_cx["B4"] = cx_date_str  # merged B4:C4 in template

            # Shopify lookups (raw export)
            addr1_lu = orders_df.groupby("Name")["Shipping Address1"].first().to_dict()
            street_lu = orders_df.groupby("Name")["Shipping Street"].first().to_dict()
            city_lu = orders_df.groupby("Name")["Shipping City"].first().to_dict()
            zip_lu = orders_df.groupby("Name")["Shipping Zip"].first().to_dict()
            prov_name_lu = orders_df.groupby("Name")["Shipping Province Name"].first().to_dict() if "Shipping Province Name" in orders_df.columns else {}
            prov_lu = orders_df.groupby("Name")["Shipping Province"].first().to_dict()
            notes_lu = orders_df.groupby("Name")["Notes"].first().to_dict()

            start_row = 6
            for row_idx, (_, r) in enumerate(cx_manifest.iterrows(), start=start_row):
                order_name = to_clean_str(r.get("D.O. No.", ""))

                inv_no = order_name
                delivery_date = cx_date_str
                store_no = ""
                store_name = clean_cell(r.get("Deliver to", ""))

                address = clean_cell(addr1_lu.get(order_name, "")) or clean_cell(street_lu.get(order_name, ""))
                suburb = clean_cell(city_lu.get(order_name, ""))

                state_full = clean_cell(prov_name_lu.get(order_name, "")) or clean_cell(prov_lu.get(order_name, ""))

                raw_post = zip_lu.get(order_name, "")
                postcode = re.sub(r"\D", "", str(raw_post))

                cartons = int(r.get("No. of Shipping Labels", 0) or 0)

                meals = float(r.get("Line Items", 0) or 0)
                weight = round(meals * 0.380, 2)

                inv_value = ""
                cod = ""
                pallets = ""
                temp = "Chilled"
                comment = clean_cell(notes_lu.get(order_name, ""))

                values = [
                    inv_no, delivery_date, store_no, store_name, address, suburb, state_full, postcode,
                    cartons, pallets, weight, inv_value, cod, temp, comment
                ]

                for col_idx, val in enumerate(values, start=1):
                    ws_cx.cell(row=row_idx, column=col_idx, value=val)

            cx_buffer = BytesIO()
            wb_cx.save(cx_buffer)
            zipf.writestr("CX_Manifest.xlsx", cx_buffer.getvalue())
        add_to_zip_excel(other_manifest, "Other_Manifest.xlsx")

        # DK Distribution (Excel now)
        if len(dk_names) > 0:
            dk_src = orders_df[orders_df["Name"].isin(dk_names)]
            dk_rows = []
            today_mel = datetime.now(ZoneInfo("Australia/Melbourne")).date()
            dk_date_str = (today_mel + timedelta(days=2)).strftime("%d/%m/%Y")

            for order_name, group in dk_src.groupby("Name", sort=False):
                order_name_clean = to_clean_str(order_name)
                mrow = manifest_df[manifest_df["D.O. No."] == order_name_clean].iloc[0]
                code = order_name_clean.upper()
                delivery_type = "Commercial" if code.startswith("CEW") else "Residential"

                ship_company = clean_cell(group["Shipping Company"].iloc[0]) if len(group) else ""
                ship_name = clean_cell(group["Shipping Name"].iloc[0]) if len(group) else ""
                location = ship_company if ship_company else ship_name

                email_vals = [clean_cell(x) for x in group["Email"].unique() if clean_cell(x)]
                email = email_vals[0] if email_vals else ""

                notes_vals = [clean_cell(x) for x in group["Notes"].unique() if clean_cell(x)]
                notes_val = notes_vals[0] if notes_vals else ""

                state_vals = [clean_cell(x) for x in group["Shipping Province"].unique() if clean_cell(x)]
                state_abbrev = state_vals[0] if state_vals else ""

                dk_rows.append({
                    "Order ID": order_name_clean,
                    "Date": dk_date_str,
                    "Time Window": "7am - 6pm",
                    "Notes": notes_val,
                    "Address 1": clean_cell(mrow["Address 1"]),
                    "Address 2": "",
                    "Address 3": "",
                    "Postal Code": to_clean_str(mrow["Postal Code"]),
                    "City": clean_cell(mrow["Address 2"]),
                    "State": state_abbrev,
                    "Country": "Australia",
                    "Location": location,
                    "Last Name": "",
                    "Phone": to_clean_str(mrow["Phone No."]),
                    "Delivery Instructions": clean_cell(mrow["Instructions"]),
                    "Email": email,
                    "DELIVERY TYPE": delivery_type,
                    "Volume": to_intish_str(mrow["No. of Shipping Labels"]),
                    "NOTES": ""
                })

            dk_df = pd.DataFrame(dk_rows)
            add_to_zip_excel(dk_df, "DK_Manifest.xlsx")

    output.seek(0)
    st.download_button(
        label="Download Manifests ZIP",
        data=output,
        file_name="CleanEats_Manifests.zip",
        mime="application/zip"
    )
